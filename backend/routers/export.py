import csv
import io
from typing import Any

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from supabase import Client

from dependencies import get_current_user, get_supabase, require_role
from services.audit_service import AuditAction, write_audit
from services.scoring import calc_score

router = APIRouter(prefix="/export", tags=["export"])

HEADERS = [
    "Employee",
    "Department",
    "Goal Title",
    "UoM Type",
    "Direction",
    "Baseline",
    "Target",
    "Planned Date",
    "Weightage",
    "Q1 Actual",
    "Q1 Score",
    "Q2 Actual",
    "Q2 Score",
    "Q3 Actual",
    "Q3 Score",
    "Q4 Actual",
    "Q4 Score",
]


def _build_rows(supabase: Client) -> list[list[Any]]:
    goals = (
        supabase.table("goals")
        .select("*, users!goals_employee_id_fkey(full_name, departments(name))")
        .eq("status", "approved")
        .execute()
        .data
        or []
    )

    goal_ids = [g["id"] for g in goals]
    actuals_map: dict[str, dict[str, dict]] = {}
    if goal_ids:
        all_actuals = (
            supabase.table("goal_actuals")
            .select("*")
            .in_("goal_id", goal_ids)
            .execute()
            .data
            or []
        )
        for a in all_actuals:
            actuals_map.setdefault(a["goal_id"], {})[a["quarter"]] = a

    rows = []
    for goal in goals:
        user_data = goal.get("users") or {}
        dept_data = user_data.get("departments") or {}
        employee_name = user_data.get("full_name", "")
        dept_name = dept_data.get("name", "") if dept_data else ""

        row: list[Any] = [
            employee_name,
            dept_name,
            goal["title"],
            goal["uom_type"],
            goal.get("direction", ""),
            goal["baseline"],
            goal["target"],
            str(goal["planned_date"]),
            goal["weightage"],
        ]

        goal_actuals = actuals_map.get(goal["id"], {})
        for quarter in ["Q1", "Q2", "Q3", "Q4"]:
            a = goal_actuals.get(quarter)
            if a:
                score = calc_score(
                    goal["uom_type"],
                    goal.get("direction"),
                    goal["baseline"],
                    goal["target"],
                    a["actual"],
                    goal.get("planned_date"),
                    a.get("actual_date"),
                )
                row.extend([a["actual"], round(score, 2)])
            else:
                row.extend([None, None])

        rows.append(row)

    return rows


@router.get("/csv")
async def export_csv(
    supabase: Client = Depends(get_supabase),
    current_user: dict = Depends(require_role("manager", "admin")),
):
    rows = _build_rows(supabase)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    writer.writerows(rows)
    output.seek(0)

    await write_audit(supabase, current_user["id"], AuditAction.EXPORT_CSV, "export")

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=goals_export.csv"},
    )


def _score_fill(score: Any) -> PatternFill:
    if score is None:
        return PatternFill()
    if score >= 90:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if score >= 70:
        return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


@router.get("/excel")
async def export_excel(
    supabase: Client = Depends(get_supabase),
    current_user: dict = Depends(require_role("manager", "admin")),
):
    rows = _build_rows(supabase)

    wb = Workbook()
    ws = wb.active
    ws.title = "Goals Export"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # score column indices (0-based in row, 1-based in worksheet)
    # Q1 Score=11, Q2 Score=13, Q3 Score=15, Q4 Score=17 (1-indexed)
    score_cols = {11, 13, 15, 17}

    for row_data in rows:
        ws.append(row_data)
        row_idx = ws.max_row
        for col_idx in score_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = _score_fill(cell.value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    await write_audit(supabase, current_user["id"], AuditAction.EXPORT_EXCEL, "export")

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=goals_export.xlsx"},
    )
